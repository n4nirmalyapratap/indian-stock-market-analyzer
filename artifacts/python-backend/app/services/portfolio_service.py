"""
Portfolio Manager Service — Postgres-backed CRUD + live valuation.

Stores per-user portfolios and holdings. Supports importing Zerodha Console
and Upstox tradebook CSVs (no broker key required at this stage).

Live valuation, day P&L, total P&L, sector / market-cap allocations and
concentration warnings are computed from our existing PriceService cache so we
never re-fetch quotes that the dashboard already has.

Tables (managed by app.lib.auth_store.ensure_primary_schema):
  portfolios               (id, user_id, name, base_currency, cash, created_at, updated_at)
  portfolio_transactions   (id, portfolio_id, symbol, side BUY|SELL|DIVIDEND,
                            qty, price, fees, traded_at, source, note, inserted_at)

Holdings are *derived* from the transactions table (weighted-avg cost) — that
way every dividend / sell / buy tweaks the position in one place and the books
always reconcile to the trade history.

Migration note: this used to be a local SQLite file
(``artifacts/python-backend/market_cache/portfolio.db``). The schema lives in
the central PG bootstrap now so multi-instance deployments don't corrupt the
SQLite write log.
"""
from __future__ import annotations

import asyncio
import csv
import difflib
import io
import logging
import re
import time
import uuid
from datetime import datetime, timezone
from typing import Any, Optional

from psycopg.rows import dict_row

import yfinance as yf

from app.lib.auth_store import ensure_primary_schema, get_conn
from app.lib.sector_utils import classify_sector, classify_market_cap

# ── yfinance .info enrichment cache ──────────────────────────────────────────
# Stores (fetched_at_epoch, {sector, industry, marketCap}) per symbol.
# TTL: 6 h — sector rarely changes, we just need it once per session.
_INFO_CACHE: dict[str, tuple[float, dict]] = {}
_INFO_TTL   = 6 * 3600


def _yf_info_enrich(symbol: str) -> dict:
    """Fetch sector/industry/marketCap from yfinance .info (synchronous).
    Results are cached for 6 hours.  Never raises — returns {} on any error.
    """
    import time as _time
    now = _time.time()
    cached = _INFO_CACHE.get(symbol)
    if cached and (now - cached[0]) < _INFO_TTL:
        return cached[1]
    try:
        suffix = ".NS" if not symbol.endswith((".NS", ".BO")) else ""
        info = yf.Ticker(symbol + suffix).info or {}
        data = {
            "sector":    info.get("sector")    or info.get("sectorDisp"),
            "industry":  info.get("industry")  or info.get("industryDisp"),
            "marketCap": info.get("marketCap") or info.get("nonDilutedMarketCap"),
        }
        _INFO_CACHE[symbol] = (now, data)
        return data
    except Exception:
        _INFO_CACHE[symbol] = (now, {})
        return {}

logger = logging.getLogger(__name__)


def _connect():
    """Open a Postgres connection with dict-row results, matching the previous
    SQLite ``conn.row_factory = sqlite3.Row`` access pattern (``row["col"]``).
    """
    conn = get_conn()
    # auth_store.get_conn already configures dict_row, but we re-set here as a
    # belt-and-braces guard against a future change in get_conn.
    conn.row_factory = dict_row  # type: ignore[attr-defined]
    return conn


# Schema is owned by auth_store.ensure_primary_schema(); call it once at
# import so a hot-restart can serve immediately without waiting for the
# lifespan to fire.
ensure_primary_schema()


# ── Helpers ──────────────────────────────────────────────────────────────────

def _now_ms() -> int:
    return int(time.time() * 1000)


def _iso(ts: int | None = None) -> str:
    return datetime.fromtimestamp((ts or _now_ms()) / 1000, tz=timezone.utc).isoformat()


def _row_to_portfolio(row: dict) -> dict:
    return {
        "id":            row["id"],
        "userId":        row["user_id"],
        "name":          row["name"],
        "baseCurrency":  row["base_currency"],
        "cash":          float(row["cash"] or 0),
        "createdAt":     _iso(row["created_at"]),
        "updatedAt":     _iso(row["updated_at"]),
    }


def _row_to_tx(row: dict) -> dict:
    return {
        "id":          row["id"],
        "portfolioId": row["portfolio_id"],
        "symbol":      row["symbol"],
        "side":        row["side"],
        "qty":         float(row["qty"]),
        "price":       float(row["price"]),
        "fees":        float(row["fees"] or 0),
        "tradedAt":    row["traded_at"],
        "source":      row["source"],
        "note":        row["note"],
    }


def _norm_symbol(sym: str) -> str:
    s = (sym or "").strip().upper()
    # Strip exchange suffixes that show up in broker statements
    for suffix in ("-EQ", ".NS", ".BO", ":NSE", ":BSE"):
        if s.endswith(suffix):
            s = s[: -len(suffix)]
    return s


# ── Portfolio CRUD ───────────────────────────────────────────────────────────

def list_portfolios(user_id: str) -> list[dict]:
    with _connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM portfolios WHERE user_id=%s ORDER BY created_at ASC",
                (user_id,),
            )
            rows = cur.fetchall()
    return [_row_to_portfolio(r) for r in rows]


def get_portfolio(user_id: str, portfolio_id: str) -> Optional[dict]:
    with _connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM portfolios WHERE id=%s AND user_id=%s",
                (portfolio_id, user_id),
            )
            row = cur.fetchone()
    return _row_to_portfolio(row) if row else None


def create_portfolio(user_id: str, name: str, cash: float = 0.0,
                     base_currency: str = "INR") -> dict:
    pid = str(uuid.uuid4())
    now = _now_ms()
    with _connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO portfolios(id,user_id,name,base_currency,cash,created_at,updated_at) "
                "VALUES(%s,%s,%s,%s,%s,%s,%s)",
                (pid, user_id, (name or "My Portfolio").strip()[:80],
                 (base_currency or "INR").upper()[:8], float(cash or 0), now, now),
            )
        conn.commit()
    return get_portfolio(user_id, pid)  # type: ignore[return-value]


def update_portfolio(user_id: str, portfolio_id: str,
                     name: Optional[str] = None,
                     cash: Optional[float] = None) -> Optional[dict]:
    p = get_portfolio(user_id, portfolio_id)
    if not p:
        return None
    fields, values = [], []
    if name is not None:
        fields.append("name=%s")
        values.append(name.strip()[:80])
    if cash is not None:
        fields.append("cash=%s")
        values.append(float(cash))
    if not fields:
        return p
    fields.append("updated_at=%s")
    values.append(_now_ms())
    values.extend([portfolio_id, user_id])
    with _connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE portfolios SET {', '.join(fields)} WHERE id=%s AND user_id=%s",
                tuple(values),
            )
        conn.commit()
    return get_portfolio(user_id, portfolio_id)


def delete_portfolio(user_id: str, portfolio_id: str) -> bool:
    with _connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM portfolios WHERE id=%s AND user_id=%s",
                (portfolio_id, user_id),
            )
            deleted = cur.rowcount > 0
        conn.commit()
    return deleted


# ── Transactions ─────────────────────────────────────────────────────────────

def list_transactions(user_id: str, portfolio_id: str,
                      symbol: Optional[str] = None) -> list[dict]:
    if not get_portfolio(user_id, portfolio_id):
        return []
    with _connect() as conn:
        with conn.cursor() as cur:
            if symbol:
                cur.execute(
                    "SELECT * FROM portfolio_transactions "
                    "WHERE portfolio_id=%s AND symbol=%s "
                    "ORDER BY traded_at DESC, inserted_at DESC",
                    (portfolio_id, _norm_symbol(symbol)),
                )
            else:
                cur.execute(
                    "SELECT * FROM portfolio_transactions "
                    "WHERE portfolio_id=%s "
                    "ORDER BY traded_at DESC, inserted_at DESC",
                    (portfolio_id,),
                )
            rows = cur.fetchall()
    return [_row_to_tx(r) for r in rows]


def add_transaction(user_id: str, portfolio_id: str, *,
                    symbol: str, side: str, qty: float, price: float,
                    fees: float = 0.0, traded_at: Optional[str] = None,
                    source: str = "manual",
                    note: Optional[str] = None) -> Optional[dict]:
    if not get_portfolio(user_id, portfolio_id):
        return None
    side_u = (side or "").upper()
    if side_u not in ("BUY", "SELL", "DIVIDEND"):
        raise ValueError(f"Invalid side: {side!r}")
    qty_f = abs(float(qty))
    price_f = float(price)
    if qty_f <= 0:
        raise ValueError("Quantity must be positive")
    if price_f < 0:
        raise ValueError("Price must be non-negative")

    tx_id = str(uuid.uuid4())
    iso = (traded_at or _iso())[:32]
    now = _now_ms()
    # Cash bookkeeping: SELL / BUY moves cash; DIVIDEND adds cash
    if side_u == "BUY":
        cash_delta = -(qty_f * price_f + (fees or 0))
    elif side_u == "SELL":
        cash_delta = +(qty_f * price_f - (fees or 0))
    else:  # DIVIDEND
        cash_delta = +(qty_f * price_f)

    with _connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO portfolio_transactions"
                "(id,portfolio_id,symbol,side,qty,price,fees,"
                " traded_at,source,note,inserted_at) "
                "VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (tx_id, portfolio_id, _norm_symbol(symbol), side_u,
                 qty_f, price_f, float(fees or 0), iso, source[:32], note, now),
            )
            cur.execute(
                "UPDATE portfolios SET cash = cash + %s, updated_at = %s "
                "WHERE id=%s",
                (cash_delta, now, portfolio_id),
            )
            cur.execute(
                "SELECT * FROM portfolio_transactions WHERE id=%s",
                (tx_id,),
            )
            row = cur.fetchone()
        conn.commit()
    return _row_to_tx(row)


def delete_transaction(user_id: str, portfolio_id: str, tx_id: str) -> bool:
    if not get_portfolio(user_id, portfolio_id):
        return False
    with _connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM portfolio_transactions "
                "WHERE id=%s AND portfolio_id=%s",
                (tx_id, portfolio_id),
            )
            row = cur.fetchone()
            if not row:
                return False

            side = row["side"]
            qty = float(row["qty"])
            price = float(row["price"])
            fees = float(row["fees"] or 0)
            if side == "BUY":
                cash_delta = +(qty * price + fees)
            elif side == "SELL":
                cash_delta = -(qty * price - fees)
            else:
                cash_delta = -(qty * price)

            cur.execute(
                "DELETE FROM portfolio_transactions WHERE id=%s",
                (tx_id,),
            )
            cur.execute(
                "UPDATE portfolios SET cash = cash + %s, updated_at = %s "
                "WHERE id=%s",
                (cash_delta, _now_ms(), portfolio_id),
            )
        conn.commit()
    return True


def delete_transactions_bulk(user_id: str, portfolio_id: str,
                              tx_ids: list[str]) -> dict:
    """Delete multiple transactions and roll back their cash impact in one
    atomic round-trip.

    Each row's cash delta is reversed — the same arithmetic as
    ``delete_transaction`` but vectorised. We do this in a single
    transaction so a mid-batch crash leaves the portfolio in a consistent
    state (either every row + cash adjustment lands, or none of them do).

    Returns ``{"deleted": int, "skipped": int}``. Rows that don't belong
    to this portfolio (e.g. a stale tx_id from a different page) are
    skipped, not errored — keeps batch DELETE forgiving.
    """
    if not get_portfolio(user_id, portfolio_id):
        return {"deleted": 0, "skipped": len(tx_ids or []), "error": "portfolio not found"}
    if not tx_ids:
        return {"deleted": 0, "skipped": 0}

    clean_ids = [str(t) for t in tx_ids if t]
    if not clean_ids:
        return {"deleted": 0, "skipped": 0}

    with _connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, side, qty, price, fees FROM portfolio_transactions "
                "WHERE id = ANY(%s) AND portfolio_id = %s",
                (clean_ids, portfolio_id),
            )
            rows = cur.fetchall()
            if not rows:
                return {"deleted": 0, "skipped": len(clean_ids)}

            total_cash_delta = 0.0
            for row in rows:
                side  = row["side"]
                qty   = float(row["qty"])
                price = float(row["price"])
                fees  = float(row["fees"] or 0)
                if side == "BUY":
                    total_cash_delta += +(qty * price + fees)
                elif side == "SELL":
                    total_cash_delta += -(qty * price - fees)
                else:  # DIVIDEND
                    total_cash_delta += -(qty * price)

            cur.execute(
                "DELETE FROM portfolio_transactions WHERE id = ANY(%s) "
                "AND portfolio_id = %s",
                (clean_ids, portfolio_id),
            )
            deleted = cur.rowcount
            cur.execute(
                "UPDATE portfolios SET cash = cash + %s, updated_at = %s "
                "WHERE id = %s",
                (total_cash_delta, _now_ms(), portfolio_id),
            )
        conn.commit()

    return {"deleted": deleted, "skipped": len(clean_ids) - deleted}


# ── Holdings derivation ──────────────────────────────────────────────────────

def derive_holdings(portfolio_id: str) -> list[dict]:
    """
    Roll up transactions into per-symbol holdings.

      qty       = sum(BUY) − sum(SELL)
      avg_cost  = weighted avg of net BUY cost  (sells reduce qty proportionally;
                  the cost basis per remaining share stays the historical avg).
      realised  = sum of (sell_price − running_avg_cost) * sell_qty
      dividends = sum of DIVIDEND.qty * DIVIDEND.price

    This is the standard Indian retail-tax book-keeping convention (FIFO is
    only enforced at tax-filing time; for live P&L we use weighted-avg cost).
    """
    with _connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM portfolio_transactions WHERE portfolio_id=%s "
                "ORDER BY traded_at ASC, inserted_at ASC",
                (portfolio_id,),
            )
            rows = cur.fetchall()

    book: dict[str, dict] = {}
    for r in rows:
        sym  = r["symbol"]
        side = r["side"]
        qty  = float(r["qty"])
        px   = float(r["price"])
        fees = float(r["fees"] or 0)

        h = book.setdefault(sym, {
            "symbol":   sym,
            "qty":      0.0,
            "avgCost":  0.0,
            "invested": 0.0,
            "realised": 0.0,
            "dividends": 0.0,
            "fees":     0.0,
            "buys":     0,
            "sells":    0,
            "firstTradedAt": r["traded_at"],
            "lastTradedAt":  r["traded_at"],
        })

        h["lastTradedAt"] = r["traded_at"]

        if side == "BUY":
            new_qty       = h["qty"] + qty
            new_invested  = h["invested"] + qty * px
            h["qty"]      = new_qty
            h["invested"] = new_invested
            h["avgCost"]  = (new_invested / new_qty) if new_qty > 0 else 0.0
            h["fees"]    += fees
            h["buys"]    += 1
        elif side == "SELL":
            avg          = h["avgCost"]
            sell_qty     = min(qty, h["qty"])  # cap at current qty
            h["realised"] += (px - avg) * sell_qty - fees
            h["qty"]      -= sell_qty
            h["invested"] = h["avgCost"] * h["qty"]   # remaining cost basis
            h["fees"]    += fees
            h["sells"]   += 1
        elif side == "DIVIDEND":
            h["dividends"] += qty * px

    # Drop fully-closed positions (qty effectively zero) but keep their P&L
    # under "closed" so the UI can show realised P&L on positions that were
    # entered + exited.
    open_holdings, closed_holdings = [], []
    for h in book.values():
        h["qty"]      = round(h["qty"], 6)
        h["avgCost"]  = round(h["avgCost"], 4)
        h["invested"] = round(h["invested"], 2)
        h["realised"] = round(h["realised"], 2)
        h["dividends"] = round(h["dividends"], 2)
        if h["qty"] > 1e-6:
            open_holdings.append(h)
        else:
            h["qty"] = 0.0
            closed_holdings.append(h)
    return open_holdings + closed_holdings


# ── Live valuation ───────────────────────────────────────────────────────────

async def value_portfolio(user_id: str, portfolio_id: str,
                          price_service) -> Optional[dict]:
    """
    Compute live valuation of a portfolio.

    Reuses our existing PriceService (NSE → Yahoo → disk cache) so we never
    duplicate quote calls that the dashboard already issued.
    """
    p = get_portfolio(user_id, portfolio_id)
    if not p:
        return None
    holdings = derive_holdings(portfolio_id)
    open_holdings = [h for h in holdings if h["qty"] > 0]

    if not open_holdings:
        return {
            "portfolio":       p,
            "holdings":        [],
            "closedHoldings":  [h for h in holdings if h["qty"] == 0],
            "totals": {
                "cash":           p["cash"],
                "marketValue":    0.0,
                "investedValue":  0.0,
                "dayPnl":         0.0,
                "dayPnlPct":      0.0,
                "unrealisedPnl":  0.0,
                "unrealisedPnlPct": 0.0,
                "realisedPnl":    sum(h["realised"] for h in holdings),
                "dividendsRcvd":  sum(h["dividends"] for h in holdings),
                "totalEquity":    p["cash"],
            },
            "concentration": [],
            "fetchedAt":     _iso(),
        }

    # Fetch all quotes concurrently via PriceService
    async def _q(sym: str) -> tuple[str, dict]:
        try:
            qm = await price_service.get_quote_with_meta(sym, cross_check=False)
            return sym, (qm or {}).get("quote") or {}
        except Exception as exc:
            logger.warning("portfolio: quote failed for %s: %s", sym, exc)
            return sym, {}

    quotes = dict(await asyncio.gather(*[_q(h["symbol"]) for h in open_holdings]))

    # Enrich quotes that have no sector/industry/marketCap from the fast chart
    # endpoint (common for micro-caps).  yfinance .info has this; we fetch it
    # concurrently and cache for 6 h so repeat requests are instant.
    needs_enrich = [
        h["symbol"] for h in open_holdings
        if not (quotes.get(h["symbol"], {}).get("sector")
                or quotes.get(h["symbol"], {}).get("industry")
                or quotes.get(h["symbol"], {}).get("marketCap"))
    ]
    if needs_enrich:
        async def _enrich_with_timeout(sym: str) -> dict:
            try:
                return await asyncio.wait_for(
                    asyncio.to_thread(_yf_info_enrich, sym),
                    timeout=5.0,
                )
            except (asyncio.TimeoutError, Exception):
                return {}

        enriched = dict(
            zip(
                needs_enrich,
                await asyncio.gather(*[_enrich_with_timeout(sym) for sym in needs_enrich]),
            )
        )
        for sym, extra in enriched.items():
            if extra:
                q = quotes.get(sym, {})
                q.setdefault("sector",    extra.get("sector"))
                q.setdefault("industry",  extra.get("industry"))
                if not q.get("marketCap") and extra.get("marketCap"):
                    q["marketCap"] = extra["marketCap"]
                quotes[sym] = q

    valued = []
    total_mv     = 0.0
    total_inv    = 0.0
    total_day    = 0.0
    total_unreal = 0.0
    sector_buckets: dict[str, float] = {}
    cap_buckets:    dict[str, float] = {}

    for h in open_holdings:
        q = quotes.get(h["symbol"], {})
        last_price = float(q.get("lastPrice") or q.get("regularMarketPrice") or 0)
        prev_close = float(q.get("previousClose") or q.get("regularMarketPreviousClose") or last_price)
        market_cap = float(q.get("marketCap") or 0)

        market_value = last_price * h["qty"]
        invested     = h["avgCost"] * h["qty"]
        unrealised   = market_value - invested
        unrealised_pct = (unrealised / invested * 100) if invested > 0 else 0.0
        day_pnl      = (last_price - prev_close) * h["qty"]
        day_pnl_pct  = ((last_price / prev_close) - 1) * 100 if prev_close > 0 else 0.0

        sector = classify_sector(
            q.get("sector") or q.get("industry"),
            symbol=h["symbol"],
        )
        sector_buckets[sector] = sector_buckets.get(sector, 0) + market_value

        cap_label = classify_market_cap(market_cap)
        cap_buckets[cap_label] = cap_buckets.get(cap_label, 0) + market_value

        total_mv     += market_value
        total_inv    += invested
        total_day    += day_pnl
        total_unreal += unrealised

        valued.append({
            **h,
            "lastPrice":      round(last_price, 2),
            "previousClose":  round(prev_close, 2),
            "marketValue":    round(market_value, 2),
            "unrealisedPnl":  round(unrealised, 2),
            "unrealisedPnlPct": round(unrealised_pct, 4),
            "dayPnl":         round(day_pnl, 2),
            "dayPnlPct":      round(day_pnl_pct, 4),
            "sector":         sector,
            "marketCap":      market_cap or None,
            "marketCapBucket": cap_label,
            "weight":         0.0,  # filled below once total_mv is known
            "companyName":    q.get("companyName") or q.get("longName") or h["symbol"],
        })

    # Fill weights & build allocation arrays
    total_equity = total_mv + p["cash"]
    for v in valued:
        v["weight"] = round((v["marketValue"] / total_mv) if total_mv > 0 else 0, 4)

    allocation_sector = [
        {"label": k, "value": round(v, 2),
         "weight": round(v / total_mv, 4) if total_mv > 0 else 0}
        for k, v in sorted(sector_buckets.items(), key=lambda kv: -kv[1])
    ]
    allocation_cap = [
        {"label": k, "value": round(v, 2),
         "weight": round(v / total_mv, 4) if total_mv > 0 else 0}
        for k, v in cap_buckets.items() if v > 0
    ]

    # Concentration warnings — any single position above 25% of equity
    concentration_warnings = [
        {"symbol": v["symbol"], "weight": v["weight"], "marketValue": v["marketValue"]}
        for v in valued if v["weight"] >= 0.25
    ]

    realised  = sum(h["realised"] for h in holdings)
    dividends = sum(h["dividends"] for h in holdings)
    total_day_pct = (total_day / (total_mv - total_day)) * 100 if (total_mv - total_day) > 0 else 0.0
    total_unreal_pct = (total_unreal / total_inv) * 100 if total_inv > 0 else 0.0

    return {
        "portfolio":       p,
        "holdings":        valued,
        "closedHoldings":  [h for h in holdings if h["qty"] == 0],
        "allocation": {
            "sector":     allocation_sector,
            "marketCap":  allocation_cap,
        },
        "totals": {
            "cash":             round(p["cash"], 2),
            "marketValue":      round(total_mv, 2),
            "investedValue":    round(total_inv, 2),
            "dayPnl":           round(total_day, 2),
            "dayPnlPct":        round(total_day_pct, 4),
            "unrealisedPnl":    round(total_unreal, 2),
            "unrealisedPnlPct": round(total_unreal_pct, 4),
            "realisedPnl":      round(realised, 2),
            "dividendsRcvd":    round(dividends, 2),
            "totalEquity":      round(total_equity, 2),
        },
        "concentration": concentration_warnings,
        "fetchedAt":     _iso(),
    }


# ── CSV Import (Zerodha Console / Upstox tradebook) ──────────────────────────

ZERODHA_HEADER_HINTS = ("trade_date", "symbol", "trade_type", "quantity", "price")
UPSTOX_HEADER_HINTS  = ("date", "scrip", "buy/sell", "quantity", "price")


def _detect_format(headers: list[str]) -> str:
    h = [str(x).strip().lower() for x in headers]
    if all(any(hint in col for col in h) for hint in ZERODHA_HEADER_HINTS):
        return "zerodha"
    if all(any(hint in col for col in h) for hint in UPSTOX_HEADER_HINTS):
        return "upstox"
    return "generic"


def _row_get(row: dict, *keys: str) -> Optional[str]:
    """Case-insensitive lookup with fallback aliases."""
    lower = {k.strip().lower(): v for k, v in row.items() if k}
    for k in keys:
        v = lower.get(k.strip().lower())
        if v is not None and str(v).strip() != "":
            return str(v).strip()
    return None


def _parse_zerodha_row(row: dict) -> Optional[dict]:
    sym  = _row_get(row, "symbol", "tradingsymbol")
    side = (_row_get(row, "trade_type", "buy_sell") or "").upper()
    qty  = _row_get(row, "quantity", "qty")
    px   = _row_get(row, "price", "trade_price")
    date = _row_get(row, "trade_date", "date")
    if not (sym and side in ("BUY", "SELL", "B", "S") and qty and px and date):
        return None
    side = "BUY" if side[0] == "B" else "SELL"
    return {
        "symbol":    sym,
        "side":      side,
        "qty":       float(qty.replace(",", "")),
        "price":     float(str(px).replace(",", "")),
        "fees":      float(_row_get(row, "brokerage", "charges", "fees") or 0),
        "tradedAt":  date,
        "source":    "zerodha-csv",
    }


def _parse_upstox_row(row: dict) -> Optional[dict]:
    sym  = _row_get(row, "scrip", "symbol", "instrument")
    side = (_row_get(row, "buy/sell", "buy_sell", "trade_type") or "").upper()
    qty  = _row_get(row, "quantity", "qty")
    px   = _row_get(row, "price", "rate")
    date = _row_get(row, "date", "trade_date", "trade date")
    if not (sym and side in ("BUY", "SELL", "B", "S") and qty and px and date):
        return None
    side = "BUY" if side[0] == "B" else "SELL"
    return {
        "symbol":    sym,
        "side":      side,
        "qty":       float(str(qty).replace(",", "")),
        "price":     float(str(px).replace(",", "")),
        "fees":      float(_row_get(row, "brokerage", "fees", "charges") or 0),
        "tradedAt":  date,
        "source":    "upstox-csv",
    }


def _parse_generic_row(row: dict) -> Optional[dict]:
    sym  = _row_get(row, "symbol", "scrip", "tradingsymbol", "instrument")
    side = (_row_get(row, "side", "trade_type", "buy/sell", "buy_sell") or "").upper()
    qty  = _row_get(row, "qty", "quantity")
    px   = _row_get(row, "price", "rate", "trade_price")
    date = _row_get(row, "date", "traded_at", "trade_date")
    if not (sym and side in ("BUY", "SELL", "B", "S", "DIVIDEND", "DIV") and qty and px and date):
        return None
    side = "DIVIDEND" if side.startswith("DIV") else ("BUY" if side[0] == "B" else "SELL")
    return {
        "symbol":   sym,
        "side":     side,
        "qty":      float(str(qty).replace(",", "")),
        "price":    float(str(px).replace(",", "")),
        "fees":     float(_row_get(row, "fees", "brokerage", "charges") or 0),
        "tradedAt": date,
        "source":   "csv",
    }


def parse_tradebook_csv(text: str) -> dict:
    """Parse a CSV string and return {format, transactions, errors}.

    Two paths:
      * **Headered tradebook** (Zerodha / Upstox / generic) — finds a header
        row by keyword and dispatches to the matching row parser.
      * **Headerless holdings export** (e.g. Dhan) — when no header is
        found, falls through to shape-based detection. Each holding becomes
        a synthetic BUY at the file's BuyAvg using today's date.
    """
    if not text or not text.strip():
        return {"format": "unknown", "transactions": [], "errors": ["Empty CSV"]}

    lines = text.splitlines()

    # ── Headered path (Zerodha / Upstox / generic) ────────────────────────────
    header_idx = _find_header_row(lines)
    if header_idx is not None:
        return _parse_headered_csv(lines[header_idx:])

    # ── Headerless path (Dhan-style holdings export) ──────────────────────────
    if _detect_headerless_format(lines) == "dhan-holdings":
        return _parse_dhan_holdings(lines, _iso())

    return {
        "format": "unknown",
        "transactions": [],
        "errors": [
            "Could not detect file format. Expected either a tradebook with "
            "headers (Zerodha/Upstox: symbol, side, qty, price, date) or a "
            "Dhan-style holdings export (Name, Qty, BuyAvg, ...)."
        ],
    }


def _find_header_row(lines: list[str]) -> Optional[int]:
    """Return the index of the first row that looks like a CSV header, or
    None if no header keyword is found. Header keywords: symbol, scrip,
    instrument, tradingsymbol — covers Zerodha/Upstox/generic tradebooks."""
    for i, ln in enumerate(lines):
        low = ln.lower()
        if "," in ln and any(
            h in low for h in ("symbol", "scrip", "instrument", "tradingsymbol")
        ):
            return i
    return None


def _parse_headered_csv(header_and_data_lines: list[str]) -> dict:
    """Parse rows that have a recognizable header line at index 0 of the
    slice (Zerodha / Upstox / generic). Extracted so the dispatch logic in
    `parse_tradebook_csv` stays linear and readable."""
    cleaned = "\n".join(header_and_data_lines)
    reader  = csv.DictReader(io.StringIO(cleaned))
    headers = reader.fieldnames or []
    fmt     = _detect_format(headers)
    parser  = {
        "zerodha": _parse_zerodha_row,
        "upstox":  _parse_upstox_row,
    }.get(fmt, _parse_generic_row)

    txs: list[dict] = []
    errors: list[str] = []
    for i, row in enumerate(reader, start=2):  # 2 = first data row in source
        try:
            parsed = parser(row)
            if parsed:
                txs.append(parsed)
        except Exception as exc:
            errors.append(f"row {i}: {exc}")
    return {"format": fmt, "transactions": txs, "errors": errors}


# ── Headerless holdings-export support (Dhan, etc.) ──────────────────────────

_HOLDINGS_FOOTER_PREFIXES = (
    "investment", "note", "total", "current value", "overall", "today's p&l",
)


def _is_number(s: str) -> bool:
    """True if `s` parses cleanly as a number once commas are stripped."""
    try:
        float((s or "").replace(",", "").strip())
        return True
    except (ValueError, AttributeError):
        return False


def _detect_headerless_format(lines: list[str]) -> Optional[str]:
    """When no header row is found, detect a headerless broker export by
    row shape. Currently recognizes Dhan portfolio CSV (8 columns:
    Name, Qty, BuyAvg, Investment, LTP, CurrentValue, P&L, P&L%).

    Scans the first 5 non-blank rows looking for ANY row whose first cell
    is non-numeric (a company name) and whose next two cells parse as
    numbers (qty + BuyAvg). One match is enough — footer rows in this
    format have a non-numeric third cell ('Current Value'), so they
    won't false-positive."""
    checked = 0
    for ln in lines:
        if not ln.strip():
            continue
        cells = [c.strip() for c in ln.split(",")]
        if len(cells) < 8:
            continue
        if cells[0] and not _is_number(cells[0]) and _is_number(cells[1]) and _is_number(cells[2]):
            return "dhan-holdings"
        checked += 1
        if checked >= 5:
            break
    return None


def _parse_dhan_holdings(lines: list[str], tradedAt_iso: str) -> dict:
    """Parse a Dhan portfolio export — each holding becomes one synthetic
    BUY transaction.

    Column layout (positional, no headers):
        0: Name       1: Qty       2: BuyAvg    3: Investment
        4: LTP        5: CurValue  6: P&L       7: P&L%

    Behaviour:
      * Blank rows and footer rows ('Investment', 'NOTE', 'Total', …)
        are silently skipped.
      * Symbol is resolved from the company name via universe.COMPANY_MAP;
        unmatched names fall through to an uppercased raw name so the row
        still imports and the user can fix the ticker manually.
      * Rows with qty=0 or invalid price are skipped with an error.
    """
    reader  = csv.reader(io.StringIO("\n".join(lines)))
    txs: list[dict]    = []
    errors: list[str] = []
    for i, cells in enumerate(reader, start=1):
        if not cells or not cells[0].strip():
            continue
        first_low = cells[0].strip().lower()
        if any(first_low.startswith(p) for p in _HOLDINGS_FOOTER_PREFIXES):
            continue
        if len(cells) < 3:
            continue
        name = cells[0].strip()
        try:
            qty     = float(cells[1].replace(",", "").strip())
            buy_avg = float(cells[2].replace(",", "").strip())
        except (ValueError, IndexError) as exc:
            errors.append(f"row {i} ({name}): {exc}")
            continue
        if qty <= 0 or buy_avg < 0:
            errors.append(f"row {i} ({name}): non-positive qty or price")
            continue
        txs.append({
            "symbol":   _resolve_symbol(name),
            "side":     "BUY",
            "qty":      qty,
            "price":    buy_avg,
            "fees":     0.0,
            "tradedAt": tradedAt_iso,
            "source":   "dhan-csv",
        })
    return {"format": "dhan-holdings", "transactions": txs, "errors": errors}


_NORMALIZE_SUFFIX_RE = re.compile(
    r"\b("
    r"limited|ltd|pvt|private|"
    r"corporation|corp|"
    r"company|co|"
    r"industries|inc|"
    r"holdings|holding|"
    r"the"
    r")\b\.?",
    re.IGNORECASE,
)
_NORMALIZE_PUNCT_RE = re.compile(r"[^a-z0-9]+", re.IGNORECASE)


def _normalize_company_name(s: str) -> str:
    """Aggressive normalization for company-name matching.

    Strips legal suffixes (LIMITED/LTD/PVT/CORP/INC/CO/HOLDINGS), drops
    all punctuation and whitespace, lowercases. The point is to collapse
    every imported variant — "HDFC Bank Ltd.", "HDFC Bank Limited",
    "HDFC BANK", "HDFCBANK" — to the same canonical key (`hdfcbank`),
    so a single equality check catches ~90% of misses.
    """
    if not s:
        return ""
    # Strip suffix words first, while spaces still exist as word
    # boundaries; then strip remaining punctuation/whitespace.
    no_suffix = _NORMALIZE_SUFFIX_RE.sub(" ", s)
    return _NORMALIZE_PUNCT_RE.sub("", no_suffix).lower()


# Lazily-built dict[normalized_name -> ticker]. Cached at module scope
# so we don't rebuild on every import row. Cleared by `_invalidate_name_map`
# if/when universe.COMPANY_MAP gets refreshed.
_NORMALIZED_NAME_MAP: Optional[dict[str, str]] = None


def _normalized_name_map() -> dict[str, str]:
    """Returns dict[normalized_name -> ticker]. Lazily constructed from
    universe.COMPANY_MAP plus the ticker symbols themselves; subsequent
    calls reuse the cached dict."""
    global _NORMALIZED_NAME_MAP
    if _NORMALIZED_NAME_MAP is not None:
        return _NORMALIZED_NAME_MAP
    from ..lib import universe  # noqa: PLC0415
    out: dict[str, str] = {}
    for sym, company in universe.COMPANY_MAP.items():
        # 1. Map the company-name key. First entry wins — universe order
        # is large-cap first, so "RELIANCE" resolves to RELIANCE rather
        # than RELIANCEPOWER or any smaller name.
        key = _normalize_company_name(company or "")
        if key and key not in out:
            out[key] = sym
        # 2. Map the ticker itself so "INFY" or "HDFCBANK" pass through
        # the same lookup even when ALL_SYMBOLS membership misses (the
        # caller may have lowercase'd or spaced the ticker).
        ticker_key = _normalize_company_name(sym)
        if ticker_key and ticker_key not in out:
            out[ticker_key] = sym
    _NORMALIZED_NAME_MAP = out
    return out


def _resolve_symbol(name: str) -> str:
    """Map a company name to an NSE ticker using `universe.COMPANY_MAP`.

    Strategy, in order:
      1. `name` is already a known ticker (case-insensitive) → use it
      2. Normalized exact match — strip "LIMITED"/"LTD"/"PVT"/"CORP"/
         "INC"/"CO"/"HOLDINGS"/punctuation/whitespace, compare canonical
         forms. Catches "HDFC Bank Ltd." == "HDFC BANK LIMITED" ==
         "HDFC Bank" == "HDFCBANK"
      3. Fuzzy match (difflib ratio ≥ 0.85) over the normalized name
         map — catches typos, vendor abbreviations ("HDFC Bnk" → HDFCBANK)
      4. Legacy prefix match — narrow case where the user supplies only
         the first 2-3 words of a long company name
      5. Legacy substring match (name ≥4 chars)
      6. Fallback: uppercased raw name (the row still imports — the
         user can fix the ticker manually; quote lookup will show '—')
    """
    name_s = (name or "").strip()
    if not name_s:
        return name_s

    # Local import: universe module imports lazily to avoid pulling in
    # the universe data when this service is loaded for tests that don't
    # exercise CSV import.
    from ..lib import universe  # noqa: PLC0415

    # 1. Already a known ticker (case-insensitive).
    name_upper = name_s.upper()
    if name_upper in universe.ALL_SYMBOLS:
        return name_upper

    # 2. Normalized exact match — covers ~90% of vendor name variants.
    name_norm = _normalize_company_name(name_s)
    if name_norm:
        nmap = _normalized_name_map()
        hit = nmap.get(name_norm)
        if hit:
            return hit

    # 3. Fuzzy match on normalized form. Threshold 0.85 is
    # high-confidence — "hdfcbnk" (typo) matches "hdfcbank" at ratio
    # ≈0.93; unrelated names score <0.5. Costs O(N) per call so we only
    # run it on misses, never on already-matched rows.
    if name_norm and len(name_norm) >= 3:
        try:
            nmap = _normalized_name_map()
            best = difflib.get_close_matches(name_norm, nmap.keys(), n=1, cutoff=0.85)
            if best:
                return nmap[best[0]]
        except Exception:
            # Fall through to legacy matchers — never break import on a
            # difflib edge case. Log at debug so the failure is grep-able
            # if a real bug ever causes the fuzzy stage to silently miss.
            logger.debug("fuzzy match error for name_norm=%r",
                         name_norm, exc_info=True)

    # 4–5. Legacy prefix/substring fallback — sometimes rescues a
    # partial name the normalizer can't reach (very short inputs like
    # "Reliance" still need to find RELIANCE).
    name_lower = name_s.lower()
    for sym, company in universe.COMPANY_MAP.items():
        if (company or "").strip().lower().startswith(name_lower):
            return sym
    if len(name_lower) >= 4:
        for sym, company in universe.COMPANY_MAP.items():
            if name_lower in (company or "").strip().lower():
                return sym

    return name_upper


# Hard limits for arbitrary-Excel uploads — protects against zip-bombs,
# multi-million-row sparse sheets, and DoS-by-decompression.  Tradebooks
# realistically contain a few thousand rows; anything beyond these caps is
# almost certainly malicious or accidentally exported with junk.
XLSX_MAX_RAW_BYTES        = 5 * 1024 * 1024     # 5 MB compressed cap
XLSX_MAX_DECOMPRESSED     = 50 * 1024 * 1024    # 50 MB uncompressed cap
XLSX_MAX_ROWS             = 50_000
XLSX_MAX_COLS             = 64
FORMULA_INJECTION_PREFIX  = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell_str(c) -> str:
    """Stringify a cell + neutralise CSV-formula-injection vectors."""
    if c is None:
        return ""
    s = str(c)
    # Excel/LibreOffice/Google Sheets treat leading =,+,-,@ as formula start.
    # Prefix with a single quote to make the cell render as text everywhere.
    if s and s[0] in FORMULA_INJECTION_PREFIX:
        return "'" + s
    return s


def xlsx_bytes_to_csv(xlsx_bytes: bytes) -> str:
    """Convert the first sheet of an .xlsx workbook to CSV text.

    Hardened against hostile uploads:
      * Compressed-size cap (XLSX_MAX_RAW_BYTES) — early reject.
      * Decompressed-size cap (XLSX_MAX_DECOMPRESSED) — defeats zip bombs.
      * Row/column caps — protects against megasheets.
      * Formula-injection neutralisation — leading =/+/-/@ is text-quoted.

    Honesty:
      * Reads only the first worksheet — multi-sheet uploads should be
        flattened by the user.
      * Empty cells become empty strings so `parse_tradebook_csv` can
        reject them with a row-level error rather than silently dropping
        the row.
    """
    import io as _io
    import csv as _csv
    import zipfile as _zip
    from openpyxl import load_workbook

    if len(xlsx_bytes) > XLSX_MAX_RAW_BYTES:
        raise ValueError(
            f"XLSX file too large ({len(xlsx_bytes):,} bytes; "
            f"limit {XLSX_MAX_RAW_BYTES:,})"
        )
    # Zip-bomb guard: openpyxl uses zipfile internally, so we can pre-check
    # the declared uncompressed sizes before handing it the full bytes.
    try:
        with _zip.ZipFile(_io.BytesIO(xlsx_bytes)) as zf:
            total_uncompressed = sum(zi.file_size for zi in zf.infolist())
            if total_uncompressed > XLSX_MAX_DECOMPRESSED:
                raise ValueError(
                    f"XLSX uncompressed payload too large "
                    f"({total_uncompressed:,} bytes; "
                    f"limit {XLSX_MAX_DECOMPRESSED:,})"
                )
    except _zip.BadZipFile as exc:
        raise ValueError(f"Not a valid .xlsx file: {exc}") from exc

    wb = load_workbook(filename=_io.BytesIO(xlsx_bytes), read_only=True,
                       data_only=True)
    ws = wb.active
    buf = _io.StringIO()
    w = _csv.writer(buf)
    rows_written = 0
    try:
        for row in ws.iter_rows(values_only=True):
            # Skip blank rows (openpyxl emits trailing all-None rows for
            # sparse sheets); keep partial rows so the parser can complain.
            if all(c is None or (isinstance(c, str) and not c.strip())
                   for c in row):
                continue
            row_clipped = row[:XLSX_MAX_COLS]
            w.writerow([_safe_cell_str(c) for c in row_clipped])
            rows_written += 1
            if rows_written > XLSX_MAX_ROWS:
                raise ValueError(
                    f"XLSX has more than {XLSX_MAX_ROWS:,} non-blank rows; "
                    "split the file before re-uploading"
                )
    finally:
        wb.close()
    return buf.getvalue()


def _commit_parsed_transactions(
    portfolio_id: str,
    txs: list[dict],
    errors: list[str],
) -> tuple[int, list[str]]:
    """Validate and atomically insert a pre-parsed transaction list.

    Extracted so both `import_transactions` (auto-detected parse) and
    `import_with_mapping` (user-specified column mapping) share the same
    validation + DB-write path. The cash rollup is single-statement so the
    portfolio's cash balance stays in lock-step with the inserted rows.

    Returns (rows_inserted, errors). The caller's errors list is mutated.
    """
    now = _now_ms()
    prepared: list[tuple] = []
    cash_delta_total = 0.0
    for tx in txs:
        try:
            side_u = (tx.get("side") or "").upper()
            if side_u not in ("BUY", "SELL", "DIVIDEND"):
                raise ValueError(f"Invalid side: {tx.get('side')!r}")
            qty_f = abs(float(tx["qty"]))
            price_f = float(tx["price"])
            fees_f = float(tx.get("fees") or 0)
            if qty_f <= 0:
                raise ValueError("Quantity must be positive")
            if price_f < 0:
                raise ValueError("Price must be non-negative")
            sym = _norm_symbol(tx["symbol"])
            iso = (tx.get("tradedAt") or _iso())[:32]
            source = (tx.get("source") or "csv")[:32]
            prepared.append((str(uuid.uuid4()), portfolio_id, sym, side_u,
                             qty_f, price_f, fees_f, iso, source, None, now))
            if side_u == "BUY":
                cash_delta_total -= qty_f * price_f + fees_f
            elif side_u == "SELL":
                cash_delta_total += qty_f * price_f - fees_f
            else:  # DIVIDEND
                cash_delta_total += qty_f * price_f
        except Exception as exc:
            errors.append(f"{tx.get('symbol')}: {exc}")

    if not prepared:
        return 0, errors
    try:
        with _connect() as conn:
            with conn.cursor() as cur:
                cur.executemany(
                    "INSERT INTO portfolio_transactions"
                    "(id,portfolio_id,symbol,side,qty,price,fees,"
                    " traded_at,source,note,inserted_at) "
                    "VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    prepared,
                )
                cur.execute(
                    "UPDATE portfolios SET cash = cash + %s, "
                    "updated_at = %s WHERE id=%s",
                    (cash_delta_total, now, portfolio_id),
                )
            conn.commit()
            return len(prepared), errors
    except Exception as exc:
        errors.append(f"import rolled back: {exc}")
        return 0, errors


def import_transactions(user_id: str, portfolio_id: str, csv_text: str) -> dict:
    """Parse a tradebook CSV and import every valid row.

    Behaviour:
      * Each row is pre-validated independently. Rows that fail validation
        are skipped and surfaced in `errors[]` so the user can fix and
        re-upload only the bad rows (a hard fail-the-whole-batch policy
        would force users to re-clean a 500-row tradebook over a single
        typo).
      * The valid rows that survive validation are then inserted in a
        single SQLite transaction (`executemany` + the rolled-up cash
        UPDATE) — so the actual DB write is atomic: either every valid
        row is committed or none of them are. There is no partial
        DB-write state on infrastructure failure.

    Returns `{format, rowsParsed, rowsInserted, errors[]}`.
    """
    if not get_portfolio(user_id, portfolio_id):
        return {"error": "portfolio not found"}

    parsed = parse_tradebook_csv(csv_text)
    inserted, parsed["errors"] = _commit_parsed_transactions(
        portfolio_id, parsed["transactions"], parsed["errors"]
    )

    return {
        "format":          parsed["format"],
        "rowsParsed":      len(parsed["transactions"]),
        "rowsInserted":    inserted,
        "errors":          parsed["errors"],
    }


# ── Two-step import: preview + apply with explicit mapping ───────────────────
#
# Lets the frontend show a 'map source columns → system fields' popup before
# committing, so users can verify the auto-detected mapping or correct it for
# a broker format we don't pre-recognize.

# System fields the mapping editor exposes. Keep this small — these are the
# only fields the importer actually uses (everything else is derived).
MAPPABLE_FIELDS = ("symbol", "side", "qty", "price", "tradedAt", "fees")


def _source_columns(lines: list[str], header_idx: Optional[int]) -> tuple[list[str], list[list[str]]]:
    """Return (column labels, sample data rows) for the mapping UI.

    For headered files the labels are the header cells verbatim. For
    headerless files we generate positional labels ('Col 1', 'Col 2', …)
    and return the first non-footer non-blank rows as samples — so the
    popup can show e.g. 'Col 1 (CRISIL)' next to 'Symbol' and the user
    can verify the mapping at a glance.
    """
    if header_idx is not None:
        reader = csv.reader(io.StringIO(lines[header_idx]))
        headers = [c.strip() for c in next(reader, [])]
        data_lines = lines[header_idx + 1:]
    else:
        headers = []
        data_lines = lines

    samples: list[list[str]] = []
    sample_reader = csv.reader(io.StringIO("\n".join(data_lines)))
    for cells in sample_reader:
        if not cells or not (cells[0] or "").strip():
            continue
        first_low = (cells[0] or "").strip().lower()
        if any(first_low.startswith(p) for p in _HOLDINGS_FOOTER_PREFIXES):
            continue
        samples.append([c.strip() for c in cells])
        if len(samples) >= 5:
            break

    if not headers and samples:
        # Headerless — synthesize positional labels matching the widest
        # sample row's column count.
        max_cols = max(len(r) for r in samples)
        headers = [f"Col {i + 1}" for i in range(max_cols)]

    return headers, samples


def _suggest_mapping(fmt: str, headers: list[str]) -> dict:
    """Pre-fill the mapping editor with our best guess.

    For headerless 'dhan-holdings' we use the known positional layout.
    For headered formats we fuzzy-match each system field to the header
    that most likely contains it (substring, case-insensitive)."""
    if fmt == "dhan-holdings":
        # Positional: Name, Qty, BuyAvg, Investment, LTP, CurValue, P&L, P&L%
        return {"symbol": 0, "qty": 1, "price": 2,
                "side": None, "tradedAt": None, "fees": None}

    cols_low = [(i, (c or "").lower()) for i, c in enumerate(headers)]

    def _find(*keywords: str) -> Optional[int]:
        for kw in keywords:
            for i, c in cols_low:
                if kw in c:
                    return i
        return None

    return {
        "symbol":   _find("symbol", "scrip", "tradingsymbol", "instrument"),
        "side":     _find("side", "trade_type", "buy/sell", "buy_sell"),
        "qty":      _find("qty", "quantity"),
        "price":    _find("price", "rate", "trade_price"),
        "tradedAt": _find("date", "trade_date", "traded_at"),
        "fees":     _find("brokerage", "fees", "charges"),
    }


def _synth_defaults(fmt: str) -> dict:
    """Default values for fields that don't exist in the source file.

    Currently only headerless holdings exports need this — they're missing
    side (we synthesize BUY) and date (we use today's ISO timestamp)."""
    if fmt == "dhan-holdings":
        return {"side": "BUY", "tradedAt": _iso()}
    return {}


def analyze_csv(csv_text: str) -> dict:
    """Inspect a CSV without committing — returns everything the frontend
    needs to render the mapping popup: detected format, source columns,
    a few sample rows, an auto-suggested mapping, and synthetic-field
    defaults. No DB writes.
    """
    if not csv_text or not csv_text.strip():
        return {
            "format": "unknown", "headerless": True,
            "sourceColumns": [], "sampleRows": [],
            "suggestedMapping": {}, "syntheticDefaults": {},
            "totalRows": 0, "errors": ["Empty CSV"],
        }
    lines = csv_text.splitlines()
    header_idx = _find_header_row(lines)
    headerless_fmt = _detect_headerless_format(lines) if header_idx is None else None

    if header_idx is not None:
        # Use a temporary DictReader to extract the format key, then walk
        # the actual rows manually below to count them.
        reader = csv.DictReader(io.StringIO("\n".join(lines[header_idx:])))
        fmt = _detect_format(reader.fieldnames or [])
    elif headerless_fmt:
        fmt = headerless_fmt
    else:
        fmt = "unknown"

    headers, samples = _source_columns(lines, header_idx)
    return {
        "format":            fmt,
        "headerless":        header_idx is None,
        "sourceColumns":     headers,
        "sampleRows":        samples,
        "suggestedMapping":  _suggest_mapping(fmt, headers),
        "syntheticDefaults": _synth_defaults(fmt),
        "totalRows":         len(samples),  # the modal only previews a few
        "errors":            [] if fmt != "unknown" else [
            "Could not auto-detect a known broker format — please map the "
            "columns manually below."
        ],
    }


def _parse_with_mapping(
    csv_text: str,
    header_idx: Optional[int],
    mapping: dict,
    synth: dict,
) -> tuple[list[dict], list[str]]:
    """Walk the CSV and build transactions using the user-provided mapping.

    `mapping` maps each system field (symbol/side/qty/price/tradedAt/fees)
    to either a column index (int) into the row's cells, or None to use
    `synth[field]` instead.
    """
    lines = csv_text.splitlines()
    data_start = (header_idx + 1) if header_idx is not None else 0
    reader = csv.reader(io.StringIO("\n".join(lines[data_start:])))
    txs: list[dict] = []
    errors: list[str] = []

    def _col(cells: list[str], field: str) -> Optional[str]:
        idx = mapping.get(field)
        if idx is None:
            v = synth.get(field)
            return None if v is None else str(v)
        if isinstance(idx, int) and 0 <= idx < len(cells):
            return cells[idx].strip()
        return None

    for i, cells in enumerate(reader, start=data_start + 1):
        if not cells or not (cells[0] or "").strip():
            continue
        first_low = (cells[0] or "").strip().lower()
        if any(first_low.startswith(p) for p in _HOLDINGS_FOOTER_PREFIXES):
            continue
        try:
            sym_raw   = _col(cells, "symbol")
            qty_raw   = _col(cells, "qty")
            price_raw = _col(cells, "price")
            side_raw  = (_col(cells, "side") or "BUY")
            date_raw  = _col(cells, "tradedAt") or _iso()
            fees_raw  = _col(cells, "fees")

            if not sym_raw or qty_raw is None or price_raw is None:
                errors.append(f"row {i}: missing symbol/qty/price")
                continue

            qty   = float(str(qty_raw).replace(",", ""))
            price = float(str(price_raw).replace(",", ""))
            fees  = float(str(fees_raw).replace(",", "")) if fees_raw else 0.0
            side  = side_raw.strip().upper()
            if side.startswith("DIV"):
                side = "DIVIDEND"
            elif side and side[0] == "S":
                side = "SELL"
            else:
                side = "BUY"  # default for B, BUY, anything else

            txs.append({
                "symbol":   _resolve_symbol(sym_raw),
                "side":     side,
                "qty":      qty,
                "price":    price,
                "fees":     fees,
                "tradedAt": date_raw,
                "source":   "mapped-csv",
            })
        except Exception as exc:
            errors.append(f"row {i}: {exc}")
    return txs, errors


def import_with_mapping(
    user_id: str,
    portfolio_id: str,
    csv_text: str,
    mapping: dict,
    synth: Optional[dict] = None,
) -> dict:
    """Import a CSV using a user-supplied column mapping (from the mapping
    popup) instead of relying on auto-detected broker format.

    Returns the same shape as `import_transactions` so the frontend doesn't
    have to branch on the response.
    """
    if not get_portfolio(user_id, portfolio_id):
        return {"error": "portfolio not found"}
    lines = csv_text.splitlines()
    header_idx = _find_header_row(lines)
    txs, errors = _parse_with_mapping(csv_text, header_idx, mapping, synth or {})
    inserted, errors = _commit_parsed_transactions(portfolio_id, txs, errors)
    return {
        "format":          "mapped",
        "rowsParsed":      len(txs),
        "rowsInserted":    inserted,
        "errors":          errors,
    }


# ── End two-step import ──────────────────────────────────────────────────────


# ── Performance vs benchmark ─────────────────────────────────────────────────

async def equity_curve(user_id: str, portfolio_id: str,
                       price_service, days: int = 365,
                       benchmark: str = "NIFTY 50") -> Optional[dict]:
    """
    Build an equity curve by replaying transactions day-by-day against the
    daily-close history of every symbol that has ever been held.

    Compares against the chosen benchmark (default NIFTY 50, normalised to
    start-of-window value of 100 for clean overlay).
    """
    p = get_portfolio(user_id, portfolio_id)
    if not p:
        return None

    txs = list_transactions(user_id, portfolio_id)
    if not txs:
        return {"portfolioId": portfolio_id, "series": [], "benchmark": benchmark}

    txs.sort(key=lambda t: t["tradedAt"])
    symbols = sorted({t["symbol"] for t in txs if t["side"] in ("BUY", "SELL")})

    # Reverse-engineer the starting cash that existed before *any* transaction
    # was applied. `p["cash"]` is the *current* book cash (after every BUY /
    # SELL / DIVIDEND has already moved it in `add_transaction`), so we
    # subtract the net cashflow of all transactions to get the t0 starting
    # cash and replay forward from there. This avoids double-counting the
    # cash deltas (which would happen if we added `p["cash"]` to a series
    # that already accumulates the same deltas).
    total_cash_delta = 0.0
    for tx in txs:
        qty   = float(tx["qty"])
        price = float(tx["price"])
        fees  = float(tx.get("fees") or 0)
        if tx["side"] == "BUY":
            total_cash_delta -= qty * price + fees
        elif tx["side"] == "SELL":
            total_cash_delta += qty * price - fees
        elif tx["side"] == "DIVIDEND":
            total_cash_delta += qty * price
    initial_cash = float(p["cash"]) - total_cash_delta

    async def _hist(sym: str) -> tuple[str, list[dict]]:
        try:
            h = await price_service.get_historical_data(sym, days)
            return sym, h or []
        except Exception as exc:
            logger.warning("portfolio.equity_curve: %s history failed: %s", sym, exc)
            return sym, []

    sym_hist = dict(await asyncio.gather(*[_hist(s) for s in symbols]))
    bench_sym, bench_data = await _hist(benchmark)

    # Build a sorted list of unique trading dates we'll project on
    all_dates = set()
    for h in sym_hist.values():
        for r in h:
            d = str(r.get("date", ""))[:10]
            if d:
                all_dates.add(d)
    for r in bench_data:
        d = str(r.get("date", ""))[:10]
        if d:
            all_dates.add(d)
    timeline = sorted(all_dates)

    # Per-symbol close lookup
    close_by_sym: dict[str, dict[str, float]] = {
        s: {str(r["date"])[:10]: float(r["close"]) for r in h if r.get("close")}
        for s, h in sym_hist.items()
    }
    bench_close = {str(r["date"])[:10]: float(r["close"]) for r in bench_data if r.get("close")}

    # Walk timeline. At each date apply transactions ≤ that date (incrementally
    # via a pointer into the sorted tx list) and value the resulting holdings.
    series: list[dict] = []
    book: dict[str, float] = {}   # symbol → qty
    cash = 0.0
    tx_iter = iter(txs)
    next_tx = next(tx_iter, None)

    prev_close: dict[str, float] = {}
    for d in timeline:
        # Apply all transactions with traded_at ≤ d
        while next_tx and str(next_tx["tradedAt"])[:10] <= d:
            sym  = next_tx["symbol"]
            side = next_tx["side"]
            qty  = float(next_tx["qty"])
            px   = float(next_tx["price"])
            fees = float(next_tx.get("fees") or 0)
            if side == "BUY":
                book[sym] = book.get(sym, 0) + qty
                cash     -= qty * px + fees
            elif side == "SELL":
                book[sym] = book.get(sym, 0) - qty
                cash     += qty * px - fees
            elif side == "DIVIDEND":
                cash     += qty * px
            next_tx = next(tx_iter, None)

        # Mark to market — fall back to previous close if we don't have a
        # quote for this date for a particular symbol
        mv = 0.0
        for sym, qty in book.items():
            if qty <= 0:
                continue
            close = close_by_sym.get(sym, {}).get(d) or prev_close.get(sym)
            if close is None:
                continue
            prev_close[sym] = close
            mv += qty * close
        # Equity = mark-to-market value of holdings + cash on that date.
        # `initial_cash` is the t0 starting cash (pre-first-tx) and `cash`
        # is the running net cash delta from replayed transactions; their
        # sum equals the cash balance as of date `d`.
        series.append({
            "date":   d,
            "equity": round(mv + initial_cash + cash, 2),
            "marketValue": round(mv, 2),
        })

    # Normalise benchmark to first portfolio equity value (so they overlay)
    bench_series = []
    if series and bench_close:
        # Find first benchmark close on/after first series date
        start_eq = None
        bench_first = None
        for pt in series:
            if pt["equity"] > 0:
                start_eq = pt["equity"]
                bench_first = bench_close.get(pt["date"]) or next(
                    (bench_close[d] for d in sorted(bench_close) if d >= pt["date"]), None)
                break
        if start_eq and bench_first:
            scale = start_eq / bench_first
            for d in timeline:
                bc = bench_close.get(d)
                if bc is not None:
                    bench_series.append({"date": d, "value": round(bc * scale, 2)})

    return {
        "portfolioId": portfolio_id,
        "series":      series,
        "benchmark":   bench_sym,
        "benchmarkSeries": bench_series,
        "fetchedAt":   _iso(),
    }
